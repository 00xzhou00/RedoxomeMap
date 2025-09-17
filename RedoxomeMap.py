import tkinter as tk
import tkinter.messagebox
from sysconfig import get_path
from tkinter import filedialog
from tkinter import ttk
import pandas as pd
import numpy as np
from PIL import Image,ImageTk
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import base64
from Picture.image import *
import os
import sys
inmzfilename=''
outfilename=''

PROJECT_DIR = os.path.dirname(__file__)
Metalioncsvfile= os.path.join(PROJECT_DIR,"Database/Metal_ion.csv")
HMDB5csvfile=os.path.join(PROJECT_DIR,"Database/HMDB5.csv")

#Get input filename
def getuploadfilename():
    infile=filedialog.askopenfilename(filetypes=(('Excel files','*.xlsx'),))
    inpath_text.set(infile)
    global inmzfilename
    inmzfilename=infile
#Get output filename
def getoutfilename():
    outfile=filedialog.askopenfilename(filetypes=(('Excel files','*.xlsx'),))
    outpath_text.set(outfile)
    global outfilename
    outfilename=outfile

#Running script
def run():
    infilename=inmzfilename
    min = float(entry_min.get())
    max = float(entry_max.get())
    error = float(entry_error.get())
    reoutybmz = []
    reoutmode=[]
    reouterror = []
    reoutaccession = []
    reoutmonisotopic_molecular_weight = []
    reoutiupac_name = []
    reoutname = []
    reoutchemical_formula = []
    reoutkegg = []

    proutybmz = []
    proutmode = []
    prouterror = []
    proutaccession = []
    proutmonisotopic_molecular_weight = []
    proutiupac_name = []
    proutname = []
    proutchemical_formula = []
    proutkegg = []

    metaloutybmz = []
    metaloutmode = []
    metalouterror = []
    metaloutaccession = []
    metaloutmonisotopic_molecular_weight = []
    metaloutiupac_name = []
    metaloutname = []
    metaloutchemical_formula = []
    metaloutkegg = []

    metaldata=pd.read_csv(Metalioncsvfile,usecols=['accession', 'monisotopic_molecular_weight', 'iupac_name', 'name','chemical_formula', 'kegg'])
    hmdbdata = pd.read_csv(HMDB5csvfile,usecols=['accession', 'monisotopic_molecular_weight', 'iupac_name', 'name','chemical_formula', 'kegg'])
    inmzdata = pd.read_excel(infilename, usecols=['m/z'])
    inmz = inmzdata['m/z'].values.tolist()
    hmdbmz = hmdbdata['monisotopic_molecular_weight'].values.tolist()
    metalmz=metaldata['monisotopic_molecular_weight'].values.tolist()
#Progress bar
    pro=tk.ttk.Progressbar(windows)
    pro.place(x=250,y=343,width=300,height=16)
    pro['maximum']=len(hmdbmz)*2
    pro['value']=0
#Cyclic matching
    for i in range(len(hmdbmz)) :
        if hmdbmz[i]<min or hmdbmz[i]>max:
            pro['value']=i
            windows.update()
            continue
        else:
            for k in range(len(inmz)):
                if hmdbmz[i]-error<inmz[k]<hmdbmz[i]+error:
                    wcc=abs(hmdbmz[i]-inmz[k])
                    reoutybmz.append(inmz[k])
                    reoutmode.append("Redox")
                    reouterror.append(wcc)
                    reoutaccession.append(hmdbdata.iloc[i,0])
                    reoutmonisotopic_molecular_weight.append(hmdbmz[i])
                    reoutiupac_name.append(hmdbdata.iloc[i,2])
                    reoutname.append(hmdbdata.iloc[i,3])
                    reoutchemical_formula.append(hmdbdata.iloc[i,4])
                    reoutkegg.append(hmdbdata.iloc[i,5])

    reoutdf = pd.DataFrame({'m/z': reoutybmz, 'Monisotopic_Mass': reoutmonisotopic_molecular_weight,'Error(Da)': reouterror, 'Mode':reoutmode,'IUPAC_Name': reoutiupac_name, 'Name': reoutname, 'Accession Number': reoutaccession,'Chemical_Formula': reoutchemical_formula, 'KEGG': reoutkegg})
    for i in range(len(metalmz)):
        for k in range(len(inmz)):
            if metalmz[i]-error<inmz[k]<metalmz[i]+error:
                wcc=abs(metalmz[i]-inmz[k])
                metaloutybmz.append(inmz[k])
                metaloutmode.append("Redox")
                metalouterror.append(wcc)
                metaloutaccession.append(metaldata.iloc[i, 0])
                metaloutmonisotopic_molecular_weight.append(metalmz[i])
                metaloutiupac_name.append(metaldata.iloc[i, 2])
                metaloutname.append(metaldata.iloc[i, 3])
                metaloutchemical_formula.append(metaldata.iloc[i, 4])
                metaloutkegg.append(metaldata.iloc[i, 5])
    metaloutdf = pd.DataFrame(
                {'m/z': metaloutybmz, 'Monisotopic_Mass': metaloutmonisotopic_molecular_weight, 'Error(Da)': metalouterror,
                 'Mode': metaloutmode, 'IUPAC_Name': metaloutiupac_name, 'Name': metaloutname,
                 'Accession Number': metaloutaccession, 'Chemical_Formula': metaloutchemical_formula, 'KEGG': metaloutkegg})

    if combobox.get()=='Negative':

        for i in range(len(hmdbmz)):
            if hmdbmz[i] < min or hmdbmz[i] > max:
                pro['value'] = i+len(hmdbmz)
                windows.update()
                continue
            else:
                for k in range(len(inmz)):
                    if hmdbmz[i] - error < inmz[k] + 1.0078 < hmdbmz[i] + error:
                        wcc = abs(hmdbmz[i] - inmz[k] - 1.0078)
                        proutybmz.append(inmz[k])
                        proutmode.append("Protonation")
                        prouterror.append(wcc)
                        proutaccession.append(hmdbdata.iloc[i, 0])
                        proutmonisotopic_molecular_weight.append(hmdbmz[i])
                        proutiupac_name.append(hmdbdata.iloc[i, 2])
                        proutname.append(hmdbdata.iloc[i, 3])
                        proutchemical_formula.append(hmdbdata.iloc[i, 4])
                        proutkegg.append(hmdbdata.iloc[i, 5])
        proutdf = pd.DataFrame({'m/z': proutybmz, 'Monisotopic_Mass': proutmonisotopic_molecular_weight, 'Error(Da)': prouterror,'Mode':proutmode,'IUPAC_Name': proutiupac_name, 'Name': proutname, 'Accession Number': proutaccession,'Chemical_Formula': proutchemical_formula, 'KEGG': proutkegg})

        pro['value'] = 0
        pro.destroy()

    elif combobox.get()=='Positive':
        for i in range(len(hmdbmz)):
            if hmdbmz[i] < min or hmdbmz[i] > max:
                pro['value'] = i + len(hmdbmz)
                windows.update()
                continue
            else:
                for k in range(len(inmz)):
                    if hmdbmz[i] - error < inmz[k] - 1.0078 < hmdbmz[i] + error:
                        wcc = abs(hmdbmz[i] - inmz[k] + 1.0078)

                        proutybmz.append(inmz[k])
                        proutmode.append("Deprotonation")
                        prouterror.append(wcc)
                        proutaccession.append(hmdbdata.iloc[i, 0])
                        proutmonisotopic_molecular_weight.append(hmdbmz[i])
                        proutiupac_name.append(hmdbdata.iloc[i, 2])
                        proutname.append(hmdbdata.iloc[i, 3])
                        proutchemical_formula.append(hmdbdata.iloc[i, 4])
                        proutkegg.append(hmdbdata.iloc[i, 5])
        proutdf = pd.DataFrame(
            {'m/z': proutybmz, 'Monisotopic_Mass': proutmonisotopic_molecular_weight, 'Error(Da)': prouterror,
             'Mode': proutmode, 'IUPAC_Name': proutiupac_name, 'Name': proutname, 'Accession Number': proutaccession,
             'Chemical_Formula': proutchemical_formula, 'KEGG': proutkegg})

        pro['value'] = 0
        pro.destroy()

#Remove duplicates
    acc1 = proutdf['Accession Number'].tolist()
    acc2 = reoutdf['Accession Number'].tolist()
    flag2 = []
    find = 0
    for i in range(len(acc2)):
        for j in range(len(acc1)):
            if acc2[i] == acc1[j]:
                find = 1
                break
        if find == 0:
            flag2.append(0)
            find = 0
        elif find == 1:
            flag2.append(1)
            find = 0

    reoutdf['find'] = flag2
    df2_dropped = reoutdf.drop(reoutdf[reoutdf['find'] == 1].index)
    df2_dropped = df2_dropped.drop(columns='find')
    ou = pd.concat([proutdf, df2_dropped], axis=0)
    ou=pd.concat([ou,metaloutdf],axis=0)
    ou.sort_values(by="m/z", inplace=True, ascending=True)
    merge_cells(ou,['m/z','Monisotopic_Mass'],outfilename)

    tk.messagebox.showinfo(title='',message='Completed successfully')

def merge_cells(df, key, output_path=None):
    """
    key De-duplicate and merge cells and center them
    Args:
        df: DataFrame input table
        key: Column name(s)
        output_path: Save Path

    Returns: Workbook

    """
    wb = Workbook()
    ws = wb.active


    col = key if isinstance(key, list) else [key]
    set_col = set(col)
    columns = [*col, *(i for i in df.columns if i not in set_col)]
    _df = df[columns]
    _df.sort_values(key, inplace=True)


    for row in dataframe_to_rows(_df, index=False, header=True):
        ws.append(row)

    align = Alignment(horizontal="center", vertical="center")
    idx = {-1, _df.shape[0] - 1}
    for i, _ in enumerate(col):
        c = _df[_].values
        idx.update(np.where(c[1:] != c[:-1])[0])
        sorted_idx = sorted(idx)
        for start, end in zip(sorted_idx[:-1], sorted_idx[1:]):

            ws.merge_cells(start_row=start + 3, end_row=end + 2, start_column=i + 1, end_column=i + 1)

            ws.cell(start + 3, i + 1).alignment = align

    if output_path:

        wb.save(output_path)

    return wb

def decode_base64(base64_string):
    decoded_bytes = base64.b64decode(base64_string)
    return decoded_bytes

def save_image(decoded_bytes, file_path):
    with open(file_path, 'wb') as image_file:
        image_file.write(decoded_bytes)

#Create a window
windows=tk.Tk()
windows.title('')
windows.geometry('804x420')


decode_bytes=decode_base64(backgroundimg_png)
file_path='backimg.png'
save_image(decode_bytes,file_path)

with Image.open(file_path) as img:
    img=img.resize((804,420))
backg=ImageTk.PhotoImage(img)
backglable=tk.Label(windows,image=backg)
backglable.pack()
os.remove('backimg.png')


text_uploadcsvfile=tk.Label(windows,text='Input',font=('Times New Roman',12))
text_uploadcsvfile.place(x=50,y=152)

uploadbutton=tk.Button(windows,text='Select file',command=getuploadfilename,width=10,height=1,font=('Times New Roman',12))
uploadbutton.pack()
uploadbutton.place(x=650,y=150)

inpath_text=tk.StringVar()
entry_inpath=tk.Entry(windows,textvariable=inpath_text,width=65,state='readonly')
entry_inpath.place(x=150,y=155)

text_Da=tk.Label(windows,text='(Da)',font=('Times New Roman',12))
text_Da.place(x=358,y=205)
text_Da2=tk.Label(windows,text='(Da)',font=('Times New Roman',12))
text_Da2.place(x=650,y=205)
text_min=tk.Label(windows,text='Mass range',font=('Times New Roman',12))
text_min.place(x=50,y=207)
text_max=tk.Label(windows,text='to',font=('Times New Roman',12))
text_max.place(x=260,y=205)
text_error=tk.Label(windows,text='Error',font=('Times New Roman',12))
text_error.place(x=500,y=207)
entry_min=tk.Entry(windows,width=5)
entry_min.pack()
entry_min.place(x=200,y=210)
entry_max=tk.Entry(windows,width=5)
entry_max.pack()
entry_max.place(x=300,y=210)
entry_error=tk.Entry(windows,width=8)
entry_error.pack()
entry_error.place(x=570,y=210)
text_outfile=tk.Label(windows,text='Results',font=('Times New Roman',12))
text_outfile.place(x=50,y=268)
outpath_text=tk.StringVar()
surebutton=tk.Button(windows,text='Save to',command=getoutfilename,width=10,height=1,font=('Times New Roman',12))
surebutton.pack()
surebutton.place(x=650,y=265)
entry4=tk.Entry(windows,textvariable=outpath_text,width=65,state='readonly')
entry4.place(x=150,y=270)

text_runmode=tk.Label(windows,text='Mode',font=('Times New Roman',12))
text_runmode.place(x=270,y=308)
combobox=ttk.Combobox(windows,values=['Negative','Positive'])
combobox.place(x=340,y=310)
combobox['state'] = "readonly"
combobox.current(0)

runbutton=tk.Button(windows,text='Run',command=run,width=10,height=1,font=('Times New Roman',12))
runbutton.pack()
runbutton.place(x=355,y=370)

windows.mainloop()